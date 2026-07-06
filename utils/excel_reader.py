from openpyxl import load_workbook


class ExcelReader:

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook[sheet_name]

    def get_row_count(self):
        return self.sheet.max_row

    def get_cell_data(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def write_result(
        self,
        row,
        actual_result,
        status,
        execution_time,
        actual_col,
        status_col,
        time_col
    ):
        self.sheet.cell(row=row, column=actual_col).value = actual_result
        self.sheet.cell(row=row, column=status_col).value = status
        self.sheet.cell(row=row, column=time_col).value = execution_time

        self.workbook.save(self.file_path)